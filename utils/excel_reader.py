#!/usr/bin/env python3
"""
Excel Reader Utility Module
Professional Contingency Analysis Suite
Reads Excel files containing contingency data with advanced error handling and validation
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Tuple, Optional, Union
import logging
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import xlrd
import warnings

warnings.filterwarnings('ignore', category=UserWarning)


class ExcelReader:
    """
    Professional Excel reader for contingency analysis data.
    Handles multiple sheet formats and provides comprehensive data extraction.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.reading_options = {
            "header_row": 0,
            "index_col": 0,
            "sheet_name": None,  # None means read all sheets
            "skip_empty_sheets": True,
            "detect_merged_cells": True,
            "preserve_formatting": True,
            "handle_formulas": True,
            "convert_to_numeric": True,
            "remove_empty_rows": True,
            "remove_empty_columns": True,
            "max_rows": None,
            "max_columns": None,
            "encoding": 'utf-8',
            "engine": 'openpyxl'  # 'openpyxl', 'xlrd', 'auto'
        }
        
        # Supported file formats
        self.supported_formats = ['.xlsx', '.xls', '.xlsm', '.xlsb']
        
        # Common contingency table indicators
        self.contingency_indicators = [
            'contingency', 'crosstab', 'cross_tab', 'frequency', 'count',
            'observed', 'expected', 'confusion', 'classification'
        ]
    
    def read_contingency_file(self, file_path: str) -> Dict[str, Any]:
        """
        Read contingency data from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary containing sheet data and metadata
        """
        try:
            self.logger.info(f"Reading Excel file: {file_path}")
            
            # Validate file
            file_validation = self._validate_file(file_path)
            if not file_validation["valid"]:
                raise ValueError(f"File validation failed: {file_validation['errors']}")
            
            # Initialize result structure
            result = {
                "file_path": file_path,
                "sheets": {},
                "metadata": {
                    "read_timestamp": datetime.now().isoformat(),
                    "file_size": Path(file_path).stat().st_size,
                    "file_format": Path(file_path).suffix.lower(),
                    "reading_options": self.reading_options.copy()
                },
                "warnings": [],
                "summary": {}
            }
            
            # Determine engine
            engine = self._determine_engine(file_path)
            
            # Read Excel file
            if engine == 'openpyxl':
                sheets_data = self._read_with_openpyxl(file_path)
            elif engine == 'xlrd':
                sheets_data = self._read_with_xlrd(file_path)
            else:
                # Try pandas default
                sheets_data = self._read_with_pandas(file_path)
            
            # Process each sheet
            for sheet_name, sheet_data in sheets_data.items():
                processed_sheet = self._process_sheet_data(sheet_data, sheet_name)
                
                if processed_sheet["valid"]:
                    result["sheets"][sheet_name] = processed_sheet["data"]
                    result["warnings"].extend(processed_sheet["warnings"])
                else:
                    result["warnings"].append(f"Sheet '{sheet_name}' could not be processed: {processed_sheet['errors']}")
            
            # Generate summary
            result["summary"] = self._generate_file_summary(result)
            
            self.logger.info(f"Successfully read {len(result['sheets'])} sheets from {file_path}")
            return result
            
        except Exception as e:
            self.logger.error(f"Failed to read Excel file '{file_path}': {str(e)}")
            raise
    
    def _validate_file(self, file_path: str) -> Dict[str, Any]:
        """Validate Excel file before reading."""
        try:
            validation_result = {
                "valid": True,
                "errors": [],
                "warnings": []
            }
            
            # Check if file exists
            if not Path(file_path).exists():
                validation_result["errors"].append(f"File does not exist: {file_path}")
                validation_result["valid"] = False
                return validation_result
            
            # Check file format
            file_suffix = Path(file_path).suffix.lower()
            if file_suffix not in self.supported_formats:
                validation_result["errors"].append(f"Unsupported file format: {file_suffix}")
                validation_result["valid"] = False
                return validation_result
            
            # Check file size
            file_size = Path(file_path).stat().st_size
            if file_size == 0:
                validation_result["errors"].append("File is empty")
                validation_result["valid"] = False
                return validation_result
            
            # Check if file is accessible
            try:
                with open(file_path, 'rb') as f:
                    f.read(1)
            except PermissionError:
                validation_result["errors"].append("File access denied")
                validation_result["valid"] = False
                return validation_result
            except Exception as e:
                validation_result["errors"].append(f"File access error: {str(e)}")
                validation_result["valid"] = False
                return validation_result
            
            # Large file warning
            if file_size > 50 * 1024 * 1024:  # 50MB
                validation_result["warnings"].append("Large file detected - processing may take time")
            
            return validation_result
            
        except Exception as e:
            return {
                "valid": False,
                "errors": [f"File validation error: {str(e)}"],
                "warnings": []
            }
    
    def _determine_engine(self, file_path: str) -> str:
        """Determine the best engine for reading the file."""
        try:
            if self.reading_options["engine"] != 'auto':
                return self.reading_options["engine"]
            
            file_suffix = Path(file_path).suffix.lower()
            
            if file_suffix in ['.xlsx', '.xlsm']:
                return 'openpyxl'
            elif file_suffix in ['.xls']:
                return 'xlrd'
            else:
                return 'openpyxl'  # Default
                
        except Exception as e:
            self.logger.error(f"Error determining engine: {str(e)}")
            return 'openpyxl'
    
    def _read_with_openpyxl(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Read Excel file using openpyxl engine."""
        try:
            sheets_data = {}
            
            # Load workbook
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheet names
            sheet_names = workbook.sheetnames
            
            # Read specific sheet or all sheets
            if self.reading_options["sheet_name"] is not None:
                if isinstance(self.reading_options["sheet_name"], str):
                    sheet_names = [self.reading_options["sheet_name"]]
                elif isinstance(self.reading_options["sheet_name"], list):
                    sheet_names = self.reading_options["sheet_name"]
            
            # Read each sheet
            for sheet_name in sheet_names:
                try:
                    # Read sheet with pandas
                    sheet_df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=self.reading_options["header_row"],
                        index_col=self.reading_options["index_col"],
                        engine='openpyxl',
                        nrows=self.reading_options["max_rows"],
                        usecols=None if self.reading_options["max_columns"] is None else range(self.reading_options["max_columns"])
                    )
                    
                    # Skip empty sheets if requested
                    if self.reading_options["skip_empty_sheets"] and sheet_df.empty:
                        continue
                    
                    sheets_data[sheet_name] = sheet_df
                    
                except Exception as e:
                    self.logger.warning(f"Could not read sheet '{sheet_name}': {str(e)}")
                    continue
            
            workbook.close()
            return sheets_data
            
        except Exception as e:
            self.logger.error(f"Error reading with openpyxl: {str(e)}")
            raise
    
    def _read_with_xlrd(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Read Excel file using xlrd engine."""
        try:
            sheets_data = {}
            
            # Read specific sheet or all sheets
            sheet_name = self.reading_options["sheet_name"]
            
            if sheet_name is None:
                # Read all sheets
                xl_file = pd.ExcelFile(file_path, engine='xlrd')
                sheet_names = xl_file.sheet_names
            else:
                sheet_names = [sheet_name] if isinstance(sheet_name, str) else sheet_name
            
            # Read each sheet
            for sheet_name in sheet_names:
                try:
                    sheet_df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=self.reading_options["header_row"],
                        index_col=self.reading_options["index_col"],
                        engine='xlrd',
                        nrows=self.reading_options["max_rows"],
                        usecols=None if self.reading_options["max_columns"] is None else range(self.reading_options["max_columns"])
                    )
                    
                    # Skip empty sheets if requested
                    if self.reading_options["skip_empty_sheets"] and sheet_df.empty:
                        continue
                    
                    sheets_data[sheet_name] = sheet_df
                    
                except Exception as e:
                    self.logger.warning(f"Could not read sheet '{sheet_name}': {str(e)}")
                    continue
            
            return sheets_data
            
        except Exception as e:
            self.logger.error(f"Error reading with xlrd: {str(e)}")
            raise
    
    def _read_with_pandas(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Read Excel file using pandas default engine."""
        try:
            sheets_data = {}
            
            # Read specific sheet or all sheets
            sheet_name = self.reading_options["sheet_name"]
            
            if sheet_name is None:
                # Read all sheets
                all_sheets = pd.read_excel(
                    file_path,
                    sheet_name=None,
                    header=self.reading_options["header_row"],
                    index_col=self.reading_options["index_col"],
                    nrows=self.reading_options["max_rows"],
                    usecols=None if self.reading_options["max_columns"] is None else range(self.reading_options["max_columns"])
                )
                
                for sheet_name, sheet_df in all_sheets.items():
                    if not (self.reading_options["skip_empty_sheets"] and sheet_df.empty):
                        sheets_data[sheet_name] = sheet_df
            else:
                # Read specific sheet(s)
                sheet_names = [sheet_name] if isinstance(sheet_name, str) else sheet_name
                
                for sn in sheet_names:
                    try:
                        sheet_df = pd.read_excel(
                            file_path,
                            sheet_name=sn,
                            header=self.reading_options["header_row"],
                            index_col=self.reading_options["index_col"],
                            nrows=self.reading_options["max_rows"],
                            usecols=None if self.reading_options["max_columns"] is None else range(self.reading_options["max_columns"])
                        )
                        
                        if not (self.reading_options["skip_empty_sheets"] and sheet_df.empty):
                            sheets_data[sn] = sheet_df
                            
                    except Exception as e:
                        self.logger.warning(f"Could not read sheet '{sn}': {str(e)}")
                        continue
            
            return sheets_data
            
        except Exception as e:
            self.logger.error(f"Error reading with pandas: {str(e)}")
            raise
    
    def _process_sheet_data(self, sheet_data: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Process individual sheet data."""
        try:
            processing_result = {
                "valid": True,
                "data": sheet_data.copy(),
                "warnings": [],
                "errors": []
            }
            
            # Clean data
            processed_data = self._clean_sheet_data(sheet_data, processing_result)
            
            # Convert flattened data to contingency table if needed
            processed_data = self._convert_flattened_to_contingency(processed_data, sheet_name)
            
            # Validate contingency data
            validation_result = self._validate_contingency_data(processed_data, sheet_name)
            processing_result["warnings"].extend(validation_result["warnings"])
            
            if validation_result["errors"]:
                processing_result["errors"].extend(validation_result["errors"])
                processing_result["valid"] = False
            
            processing_result["data"] = processed_data
            
            return processing_result
            
        except Exception as e:
            return {
                "valid": False,
                "data": sheet_data,
                "warnings": [],
                "errors": [f"Sheet processing error: {str(e)}"]
            }
    
    def _clean_sheet_data(self, sheet_data: pd.DataFrame, processing_result: Dict[str, Any]) -> pd.DataFrame:
        """Clean and prepare sheet data."""
        try:
            cleaned_data = sheet_data.copy()
            
            # Remove completely empty rows
            if self.reading_options["remove_empty_rows"]:
                before_rows = len(cleaned_data)
                cleaned_data = cleaned_data.dropna(how='all')
                after_rows = len(cleaned_data)
                
                if before_rows != after_rows:
                    processing_result["warnings"].append(f"Removed {before_rows - after_rows} empty rows")
            
            # Remove completely empty columns
            if self.reading_options["remove_empty_columns"]:
                before_cols = len(cleaned_data.columns)
                cleaned_data = cleaned_data.dropna(axis=1, how='all')
                after_cols = len(cleaned_data.columns)
                
                if before_cols != after_cols:
                    processing_result["warnings"].append(f"Removed {before_cols - after_cols} empty columns")
            
            # Convert to numeric where possible
            if self.reading_options["convert_to_numeric"]:
                numeric_columns = []
                for col in cleaned_data.columns:
                    # Try to convert to numeric
                    try:
                        cleaned_data[col] = pd.to_numeric(cleaned_data[col], errors='coerce')
                        numeric_columns.append(col)
                    except:
                        # Keep original if conversion fails
                        continue
                
                if numeric_columns:
                    processing_result["warnings"].append(f"Converted {len(numeric_columns)} columns to numeric")
            
            # Handle index
            if cleaned_data.index.name is None and len(cleaned_data.index) > 0:
                # Check if index looks like row labels
                if isinstance(cleaned_data.index[0], str):
                    processing_result["warnings"].append("Detected string index - treating as row labels")
                elif np.issubdtype(cleaned_data.index.dtype, np.integer):
                    # Reset numeric index to default
                    cleaned_data.reset_index(drop=True, inplace=True)
            
            return cleaned_data
            
        except Exception as e:
            self.logger.error(f"Error cleaning sheet data: {str(e)}")
            return sheet_data
    
    def _validate_contingency_data(self, sheet_data: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Validate sheet data for contingency analysis."""
        try:
            validation_result = {
                "warnings": [],
                "errors": []
            }
            
            # Check minimum dimensions
            if sheet_data.shape[0] < 2:
                validation_result["errors"].append(f"Sheet '{sheet_name}' has fewer than 2 rows")
            
            if sheet_data.shape[1] < 2:
                validation_result["errors"].append(f"Sheet '{sheet_name}' has fewer than 2 columns")
            
            # Check for numeric data
            numeric_columns = sheet_data.select_dtypes(include=[np.number]).columns
            if len(numeric_columns) == 0:
                validation_result["errors"].append(f"Sheet '{sheet_name}' contains no numeric data")
            elif len(numeric_columns) < sheet_data.shape[1]:
                validation_result["warnings"].append(f"Sheet '{sheet_name}' has some non-numeric columns")
            
            # Check for negative values
            numeric_data = sheet_data.select_dtypes(include=[np.number])
            if (numeric_data < 0).any().any():
                validation_result["warnings"].append(f"Sheet '{sheet_name}' contains negative values")
            
            # Check for extremely large values
            if (numeric_data > 1e6).any().any():
                validation_result["warnings"].append(f"Sheet '{sheet_name}' contains very large values")
            
            # Check for potential contingency table structure
            if self._looks_like_contingency_table(sheet_data):
                validation_result["warnings"].append(f"Sheet '{sheet_name}' appears to be a contingency table")
            
            return validation_result
            
        except Exception as e:
            return {
                "warnings": [],
                "errors": [f"Validation error: {str(e)}"]
            }
    
    def _looks_like_contingency_table(self, sheet_data: pd.DataFrame) -> bool:
        """Check if sheet looks like a contingency table."""
        try:
            # Check sheet name for indicators
            sheet_name_lower = str(sheet_data.index.name).lower() if sheet_data.index.name else ""
            
            for indicator in self.contingency_indicators:
                if indicator in sheet_name_lower:
                    return True
            
            # Check if data structure looks like contingency table
            numeric_data = sheet_data.select_dtypes(include=[np.number])
            
            # Square or rectangular numeric matrix
            if numeric_data.shape[0] >= 2 and numeric_data.shape[1] >= 2:
                # Check if values are mostly integers (typical for count data)
                if numeric_data.dtypes.apply(lambda x: np.issubdtype(x, np.integer)).all():
                    return True
                
                # Check if values are non-negative (typical for frequencies)
                if (numeric_data >= 0).all().all():
                    return True
            
            return False
            
        except Exception as e:
            self.logger.error(f"Error checking contingency table structure: {str(e)}")
            return False
    
    def _convert_flattened_to_contingency(self, sheet_data: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """Convert flattened frequency data to contingency table format."""
        try:
            # Check if this looks like flattened data (many rows, few columns)
            if sheet_data.shape[0] > sheet_data.shape[1] * 2:
                self.logger.info(f"Converting flattened data in sheet '{sheet_name}' to contingency table")
                
                # Get numeric columns (excluding index-like columns)
                numeric_cols = sheet_data.select_dtypes(include=[np.number]).columns
                
                if len(numeric_cols) >= 2:
                    # Create contingency table by summing frequencies
                    contingency_data = sheet_data[numeric_cols].sum()
                    
                    # Create a square matrix (assuming equal categories)
                    n_categories = len(numeric_cols)
                    contingency_matrix = np.zeros((n_categories, n_categories))
                    
                    # Fill diagonal with category totals
                    for i, col in enumerate(numeric_cols):
                        contingency_matrix[i, i] = contingency_data[col]
                    
                    # Create DataFrame with proper labels
                    contingency_df = pd.DataFrame(
                        contingency_matrix,
                        index=numeric_cols,
                        columns=numeric_cols
                    )
                    
                    self.logger.info(f"Converted {sheet_name} to {contingency_df.shape} contingency table")
                    return contingency_df
            
            # If not flattened, return original
            return sheet_data
            
        except Exception as e:
            self.logger.error(f"Error converting flattened data: {str(e)}")
            return sheet_data
    
    def _generate_file_summary(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Generate summary of the Excel file reading process."""
        try:
            summary = {
                "total_sheets": len(result["sheets"]),
                "successful_reads": len(result["sheets"]),
                "total_warnings": len(result["warnings"]),
                "file_info": {
                    "size_mb": result["metadata"]["file_size"] / (1024 * 1024),
                    "format": result["metadata"]["file_format"]
                },
                "sheet_details": {}
            }
            
            # Analyze each sheet
            for sheet_name, sheet_data in result["sheets"].items():
                sheet_summary = {
                    "dimensions": sheet_data.shape,
                    "numeric_columns": len(sheet_data.select_dtypes(include=[np.number]).columns),
                    "total_cells": sheet_data.shape[0] * sheet_data.shape[1],
                    "non_null_cells": sheet_data.notna().sum().sum(),
                    "data_density": (sheet_data.notna().sum().sum() / (sheet_data.shape[0] * sheet_data.shape[1])) * 100
                }
                
                # Check for potential issues
                if sheet_summary["data_density"] < 50:
                    sheet_summary["quality_flag"] = "sparse_data"
                elif sheet_summary["numeric_columns"] == 0:
                    sheet_summary["quality_flag"] = "no_numeric_data"
                else:
                    sheet_summary["quality_flag"] = "good"
                
                summary["sheet_details"][sheet_name] = sheet_summary
            
            return summary
            
        except Exception as e:
            self.logger.error(f"Error generating file summary: {str(e)}")
            return {"error": str(e)}
    
    def read_specific_range(self, file_path: str, sheet_name: str, 
                           start_row: int, end_row: int, 
                           start_col: int, end_col: int) -> pd.DataFrame:
        """Read a specific range from Excel sheet."""
        try:
            # Use openpyxl to read specific range
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook[sheet_name]
            
            # Get cell range
            cell_range = worksheet.iter_rows(
                min_row=start_row, max_row=end_row,
                min_col=start_col, max_col=end_col,
                values_only=True
            )
            
            # Convert to DataFrame
            data = list(cell_range)
            df = pd.DataFrame(data)
            
            workbook.close()
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error reading specific range: {str(e)}")
            raise
    
    def get_sheet_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about sheets in Excel file without reading data."""
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            sheet_info = {
                "sheet_names": workbook.sheetnames,
                "sheet_count": len(workbook.sheetnames),
                "sheet_details": {}
            }
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Get sheet dimensions
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                sheet_info["sheet_details"][sheet_name] = {
                    "max_row": max_row,
                    "max_column": max_col,
                    "dimensions": (max_row, max_col)
                }
            
            workbook.close()
            
            return sheet_info
            
        except Exception as e:
            self.logger.error(f"Error getting sheet info: {str(e)}")
            raise
    
    def get_reading_options(self) -> Dict[str, Any]:
        """Get current reading options."""
        return self.reading_options.copy()
    
    def set_reading_options(self, options: Dict[str, Any]) -> None:
        """Set reading options."""
        self.reading_options.update(options)
        self.logger.info(f"Reading options updated: {options}")
    
    def detect_contingency_sheets(self, file_path: str) -> List[str]:
        """Detect sheets that likely contain contingency data."""
        try:
            sheet_info = self.get_sheet_info(file_path)
            contingency_sheets = []
            
            for sheet_name in sheet_info["sheet_names"]:
                # Check sheet name for indicators
                sheet_name_lower = sheet_name.lower()
                
                for indicator in self.contingency_indicators:
                    if indicator in sheet_name_lower:
                        contingency_sheets.append(sheet_name)
                        break
                else:
                    # Check sheet dimensions (contingency tables are typically square or rectangular)
                    dimensions = sheet_info["sheet_details"][sheet_name]["dimensions"]
                    if 2 <= dimensions[0] <= 20 and 2 <= dimensions[1] <= 20:
                        contingency_sheets.append(sheet_name)
            
            return contingency_sheets
            
        except Exception as e:
            self.logger.error(f"Error detecting contingency sheets: {str(e)}")
            return []
